"""Unit tests for evidentia_core.tprm.questionnaire (v0.7.9 P0.2)."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

import pytest
from evidentia_core.models.tprm import (
    CriticalityTier,
    FourthParty,
    RegulatoryClassification,
    Vendor,
    VendorType,
)
from evidentia_core.tprm.questionnaire import (
    Questionnaire,
    QuestionnaireFormat,
    generate_questionnaire,
    render_csv_questionnaire,
    shipped_formats,
)


def _make_vendor(
    name: str = "Acme Cloud",
    type_: VendorType = VendorType.SAAS,
    criticality_tier: CriticalityTier = CriticalityTier.HIGH,
    region: str | None = "us-east-1",
    fourth_parties: list[FourthParty] | None = None,
    regulatory_classification: list[RegulatoryClassification] | None = None,
) -> Vendor:
    return Vendor(
        name=name,
        type=type_,
        criticality_tier=criticality_tier,
        relationship_owner="x@x.com",
        contract_start_date=date(2025, 1, 1),
        region=region,
        fourth_parties=fourth_parties or [],
        regulatory_classification=regulatory_classification or [],
    )


# ── shipped_formats helper ─────────────────────────────────────────


class TestShippedFormats:
    def test_includes_packaged_formats(self) -> None:
        formats = shipped_formats()
        # v0.7.9 P0.2 first slice: caiq-lite + evidentia-generic.
        # P0.2 second slice adds caiq-full. sig + sig-lite are
        # BYO-template only (not packaged content).
        assert QuestionnaireFormat.CAIQ_LITE in formats
        assert QuestionnaireFormat.EVIDENTIA_GENERIC in formats
        assert QuestionnaireFormat.CAIQ_FULL in formats
        assert QuestionnaireFormat.SIG not in formats
        assert QuestionnaireFormat.SIG_LITE not in formats


# ── generate happy path ────────────────────────────────────────────


class TestGenerateQuestionnaire:
    def test_evidentia_generic_loads(self) -> None:
        v = _make_vendor()
        q = generate_questionnaire(v, QuestionnaireFormat.EVIDENTIA_GENERIC)
        assert q.format == QuestionnaireFormat.EVIDENTIA_GENERIC.value
        assert "Evidentia Generic" in q.title
        assert q.vendor.vendor_name == "Acme Cloud"
        assert q.vendor.region == "us-east-1"
        assert len(q.questions) >= 15  # ~20 questions ship today
        # Each question carries id + domain + question_text
        for question in q.questions:
            assert question.id
            assert question.domain
            assert question.question_text

    def test_caiq_lite_loads_with_attribution(self) -> None:
        v = _make_vendor()
        q = generate_questionnaire(v, QuestionnaireFormat.CAIQ_LITE)
        assert q.format == QuestionnaireFormat.CAIQ_LITE.value
        assert "CAIQ" in q.title
        # CSA CC BY 4.0 attribution required by the source license
        assert q.licensing_attribution is not None
        assert "CSA" in q.licensing_attribution or "Cloud Security Alliance" in q.licensing_attribution
        assert "CC BY 4.0" in q.licensing_attribution
        assert len(q.questions) >= 20

    def test_evidentia_generic_no_attribution(self) -> None:
        # Apache-2.0 / no third-party license; no required attribution
        v = _make_vendor()
        q = generate_questionnaire(v, QuestionnaireFormat.EVIDENTIA_GENERIC)
        assert q.licensing_attribution is None

    def test_sig_format_raises_not_implemented(self) -> None:
        v = _make_vendor()
        with pytest.raises(NotImplementedError, match="paywalled"):
            generate_questionnaire(v, QuestionnaireFormat.SIG)

    def test_sig_lite_format_raises_not_implemented(self) -> None:
        v = _make_vendor()
        with pytest.raises(NotImplementedError, match="paywalled"):
            generate_questionnaire(v, QuestionnaireFormat.SIG_LITE)


# ── prefill projection ────────────────────────────────────────────


class TestVendorPreFill:
    def test_prefill_carries_4th_parties(self) -> None:
        v = _make_vendor(
            fourth_parties=[
                FourthParty(
                    name="AWS",
                    type=VendorType.CLOUD_PROVIDER,
                    relationship="iaas",
                ),
                FourthParty(
                    name="Stripe",
                    type=VendorType.SAAS,
                    relationship="payments",
                ),
            ],
        )
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        assert q.vendor.fourth_party_count == 2
        assert "AWS" in q.vendor.fourth_party_names
        assert "Stripe" in q.vendor.fourth_party_names

    def test_prefill_carries_regulatory_classification(self) -> None:
        v = _make_vendor(
            regulatory_classification=[
                RegulatoryClassification.MODEL,
                RegulatoryClassification.CRITICAL_THIRD_PARTY,
            ],
        )
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        assert (
            RegulatoryClassification.MODEL.value
            in q.vendor.regulatory_classification
        )
        assert (
            RegulatoryClassification.CRITICAL_THIRD_PARTY.value
            in q.vendor.regulatory_classification
        )

    def test_prefill_handles_no_region(self) -> None:
        v = _make_vendor(region=None)
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        assert q.vendor.region is None

    def test_prefill_handles_indefinite_contract(self) -> None:
        v = _make_vendor()
        # contract_end_date defaults to None
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        assert q.vendor.contract_end_date is None


# ── data-file integrity ───────────────────────────────────────────


class TestDataFileIntegrity:
    """Catch data-file regressions: missing keys, dupe IDs, etc."""

    @pytest.mark.parametrize(
        "fmt",
        [
            QuestionnaireFormat.EVIDENTIA_GENERIC,
            QuestionnaireFormat.CAIQ_LITE,
        ],
    )
    def test_question_ids_unique(self, fmt: QuestionnaireFormat) -> None:
        v = _make_vendor()
        q = generate_questionnaire(v, fmt)
        ids = [question.id for question in q.questions]
        assert len(ids) == len(set(ids)), (
            f"Duplicate question IDs in {fmt.value}: "
            f"{[i for i in set(ids) if ids.count(i) > 1]}"
        )

    @pytest.mark.parametrize(
        "fmt",
        [
            QuestionnaireFormat.EVIDENTIA_GENERIC,
            QuestionnaireFormat.CAIQ_LITE,
        ],
    )
    def test_all_questions_carry_required_fields(
        self, fmt: QuestionnaireFormat
    ) -> None:
        v = _make_vendor()
        q = generate_questionnaire(v, fmt)
        for question in q.questions:
            assert question.id
            assert question.domain
            assert question.question_text
            assert isinstance(question.response_options, list)


# ── CSV rendering ──────────────────────────────────────────────────


class TestRenderCsvQuestionnaire:
    def test_csv_includes_prefill_header_section(self) -> None:
        v = _make_vendor(
            fourth_parties=[
                FourthParty(
                    name="AWS",
                    type=VendorType.CLOUD_PROVIDER,
                    relationship="iaas",
                ),
            ],
        )
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        csv_str = render_csv_questionnaire(q)
        # Header section uses # comment-prefixed lines for vendor metadata
        assert "# Vendor DD Questionnaire" in csv_str
        assert "# Vendor name" in csv_str
        assert "Acme Cloud" in csv_str
        # 4th-party block
        assert "# 4th parties" in csv_str
        assert "AWS" in csv_str

    def test_csv_includes_question_rows_with_blank_response(self) -> None:
        v = _make_vendor()
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        csv_str = render_csv_questionnaire(q)
        # Header row for question columns
        assert (
            "id,domain,question_text,response_options,notes,vendor_response"
            in csv_str
        )
        # Last column should be blank for vendor to fill
        # (every question row ends with the empty vendor_response cell)
        for line in csv_str.split("\n"):
            if line.startswith("EVG-"):
                # Quick sanity: line ends in trailing comma + nothing
                # (csv.writer emits empty-string cells as bare commas)
                assert line.rstrip("\r").endswith(",")

    def test_csv_includes_caiq_attribution(self) -> None:
        v = _make_vendor()
        q = generate_questionnaire(v, QuestionnaireFormat.CAIQ_LITE)
        csv_str = render_csv_questionnaire(q)
        assert "# Attribution" in csv_str
        assert "CSA" in csv_str or "Cloud Security Alliance" in csv_str


# ── JSON round-trip ────────────────────────────────────────────────


class TestCsvInjectionDefense:
    """H-1 Continuous-review regression — questionnaire CSV is
    explicitly designed to be sent to the vendor; vendor-controlled
    cells in the prefill header (vendor name / 4P name / region /
    relationship_owner) must not interpret as Excel formulas."""

    @pytest.mark.parametrize(
        "malicious_name",
        [
            "=cmd|'/c calc'!A0",
            "+SUM(A1)",
            "@HYPERLINK(\"http://attacker\")",
            "-formula",
        ],
    )
    def test_questionnaire_csv_neutralizes_vendor_name_formulas(
        self, malicious_name: str
    ) -> None:
        v = _make_vendor(name=malicious_name)
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        csv_str = render_csv_questionnaire(q)
        # Vendor-name cell should be prefixed with single-quote
        for line in csv_str.split("\n"):
            if line.startswith("# Vendor name,"):
                # The cell after the comma should start with "'"
                cell = line.split(",", 1)[1].strip().strip('"')
                assert cell.startswith("'"), (
                    f"Vendor name cell should be defused: {line!r}"
                )

    def test_questionnaire_csv_4p_list_cell_safe_via_count_prefix(
        self,
    ) -> None:
        # The 4P-list cell in the prefill header is wrapped as
        # ``"<count>: <name1>, <name2>"``, so it always starts with
        # a digit (the count) — Excel never interprets it as a
        # formula even when 4P names contain `=` / `+` / `-` / `@`.
        # _csv_safe additionally prefixes if the wrapped string
        # somehow leads with a formula char (defense in depth).
        v = _make_vendor(
            fourth_parties=[
                FourthParty(
                    name="=BAD()",
                    type=VendorType.SAAS,
                    relationship="malicious",
                ),
            ],
        )
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        csv_str = render_csv_questionnaire(q)
        # Find the "# 4th parties" row + verify the cell starts
        # with a digit (count), not `=`
        for line in csv_str.split("\n"):
            if line.startswith("# 4th parties,"):
                cell = line.split(",", 1)[1].strip().strip('"')
                assert cell[0].isdigit(), (
                    f"4P list cell should lead with count digit, "
                    f"not formula char: {line!r}"
                )
                # Malicious payload still appears (we don't drop it)
                assert "=BAD()" in cell or "BAD()" in cell
                break
        else:
            pytest.fail(
                f"No '# 4th parties' row in CSV:\n{csv_str}"
            )


class TestFormatStringSafety:
    """H-3 Continuous-review regression — vendor.name with `{...}`
    placeholders should not raise KeyError or attempt to substitute.
    Closes the foot-gun where a future template addition could be
    exploited via attacker-controlled vendor name."""

    def test_vendor_name_with_brace_placeholders_renders_literally(
        self,
    ) -> None:
        # Vendor name containing `.format()`-style placeholder syntax
        v = _make_vendor(name="{vendor_name}")
        # Must not raise KeyError or recursively substitute
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        # The literal braces survive into the title
        assert "{vendor_name}" in q.title

    def test_vendor_name_with_unknown_placeholder_safe(self) -> None:
        # Hypothetical attacker-supplied template-walker
        v = _make_vendor(name="{0.__class__.__init__.__globals__}")
        # Must not raise + must not actually walk attributes
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        assert "{0.__class__.__init__.__globals__}" in q.title


class TestQuestionnaireJsonRoundTrip:
    def test_round_trip(self) -> None:
        v = _make_vendor()
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        data = q.model_dump(mode="json")
        restored = Questionnaire.model_validate(data)
        assert restored.id == q.id
        assert restored.format == q.format
        assert len(restored.questions) == len(q.questions)
        assert restored.vendor.vendor_name == q.vendor.vendor_name


# ── v0.7.9 P0.2 second slice: caiq-full ────────────────────────────


class TestCaiqFull:
    def test_caiq_full_loads(self) -> None:
        v = _make_vendor()
        q = generate_questionnaire(v, QuestionnaireFormat.CAIQ_FULL)
        assert q.format == QuestionnaireFormat.CAIQ_FULL.value
        # caiq-full ships ~50 questions vs caiq-lite's ~25
        assert len(q.questions) >= 40
        # All standard CAIQ domains covered
        domains = {qq.domain for qq in q.questions}
        assert "Audit Assurance & Compliance" in domains
        assert "Identity & Access Management" in domains
        assert "Encryption & Key Management" in domains

    def test_caiq_full_carries_csa_attribution(self) -> None:
        v = _make_vendor()
        q = generate_questionnaire(v, QuestionnaireFormat.CAIQ_FULL)
        # CC BY 4.0 attribution required by CSA license
        assert q.licensing_attribution is not None
        assert "CC BY 4.0" in q.licensing_attribution
        assert "Cloud Security Alliance" in q.licensing_attribution


# ── v0.7.9 P0.2 second slice: XLSX render ──────────────────────────


class TestRenderXlsxQuestionnaire:
    def test_xlsx_round_trip_returns_bytes(self) -> None:
        from evidentia_core.tprm.questionnaire import (
            render_xlsx_questionnaire,
        )

        v = _make_vendor()
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        xlsx = render_xlsx_questionnaire(q)
        assert isinstance(xlsx, bytes)
        # XLSX files start with PK (zip magic)
        assert xlsx[:2] == b"PK"

    def test_xlsx_contains_vendor_metadata_sheet(self) -> None:
        import io

        import openpyxl
        from evidentia_core.tprm.questionnaire import (
            render_xlsx_questionnaire,
        )

        v = _make_vendor(name="ParseMe Co")
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        xlsx = render_xlsx_questionnaire(q)
        wb = openpyxl.load_workbook(filename=io.BytesIO(xlsx))
        assert "Vendor metadata" in wb.sheetnames
        meta_ws = wb["Vendor metadata"]
        meta_rows = list(meta_ws.iter_rows(values_only=True))
        # At least the standard prefill fields land here
        labels = {row[0] for row in meta_rows if row and row[0]}
        assert "Title" in labels
        assert "Vendor name" in labels
        assert "Vendor ID" in labels
        # Vendor name cell is correctly written
        for label, value, *_ in meta_rows:
            if label == "Vendor name":
                assert value == "ParseMe Co"
                break

    def test_xlsx_groups_questions_by_domain(self) -> None:
        import io

        import openpyxl
        from evidentia_core.tprm.questionnaire import (
            render_xlsx_questionnaire,
        )

        v = _make_vendor()
        q = generate_questionnaire(v, QuestionnaireFormat.CAIQ_LITE)
        xlsx = render_xlsx_questionnaire(q)
        wb = openpyxl.load_workbook(filename=io.BytesIO(xlsx))
        # One sheet per domain plus the metadata sheet
        non_meta = [s for s in wb.sheetnames if s != "Vendor metadata"]
        domains = {qq.domain for qq in q.questions}
        # Each unique domain has at least one corresponding sheet
        # (sanitized — Excel limits sheet names to 31 chars)
        assert len(non_meta) >= len(domains) - 2  # allow truncation


# ── v0.7.9 P0.2 second slice: ingest path ──────────────────────────


class TestParseCompletedQuestionnaire:
    def test_parses_json_round_trip(self, tmp_path: Path) -> None:
        from evidentia_core.tprm.questionnaire import (
            parse_completed_questionnaire,
        )

        v = _make_vendor()
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        # Simulate vendor responses by populating vendor_response on
        # the model dump
        dump = q.model_dump(mode="json")
        for question in dump["questions"]:
            question["vendor_response"] = "Yes"
        json_path = tmp_path / "completed.json"
        import json as _json

        json_path.write_text(_json.dumps(dump), encoding="utf-8")
        completed = parse_completed_questionnaire(json_path)
        assert completed.questionnaire_id == q.id
        assert completed.vendor_id == v.id
        assert len(completed.responses) == len(q.questions)
        assert all(r == "Yes" for r in completed.responses.values())

    def test_parses_csv(self, tmp_path: Path) -> None:
        from evidentia_core.tprm.questionnaire import (
            parse_completed_questionnaire,
            render_csv_questionnaire,
        )

        v = _make_vendor()
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        csv_text = render_csv_questionnaire(q)
        # Simulate one filled-in response by editing the last column
        # of the first question row (column index 5)
        lines = csv_text.splitlines()
        # Find first question row (after header row "id,domain,...")
        for i, line in enumerate(lines):
            if line.startswith("id,"):
                # Next line is first question
                target = i + 1
                cols = lines[target].split(",")
                while len(cols) < 6:
                    cols.append("")
                cols[5] = "Yes"
                lines[target] = ",".join(cols)
                break
        csv_path = tmp_path / "completed.csv"
        csv_path.write_text("\n".join(lines), encoding="utf-8")
        completed = parse_completed_questionnaire(csv_path)
        assert completed.questionnaire_id == q.id
        assert completed.vendor_id == v.id
        # At least one response captured
        assert any(r == "Yes" for r in completed.responses.values())

    def test_parses_xlsx(self, tmp_path: Path) -> None:
        import openpyxl
        from evidentia_core.tprm.questionnaire import (
            parse_completed_questionnaire,
            render_xlsx_questionnaire,
        )

        v = _make_vendor()
        q = generate_questionnaire(
            v, QuestionnaireFormat.EVIDENTIA_GENERIC
        )
        xlsx_bytes = render_xlsx_questionnaire(q)
        xlsx_path = tmp_path / "completed.xlsx"
        xlsx_path.write_bytes(xlsx_bytes)
        # Open + populate vendor_response cells in one domain sheet
        wb = openpyxl.load_workbook(filename=str(xlsx_path))
        first_domain_sheet = next(
            s for s in wb.sheetnames if s != "Vendor metadata"
        )
        ws = wb[first_domain_sheet]
        # Header row exists; row 2 is first question
        if ws.max_row >= 2:
            # vendor_response is column 5 (1-indexed)
            ws.cell(row=2, column=5).value = "Yes"
        wb.save(str(xlsx_path))
        completed = parse_completed_questionnaire(xlsx_path)
        assert completed.questionnaire_id == q.id
        assert completed.vendor_id == v.id
        # At least one response captured
        assert any(r == "Yes" for r in completed.responses.values())

    def test_unsupported_extension_errors(self, tmp_path: Path) -> None:
        from evidentia_core.tprm.questionnaire import (
            parse_completed_questionnaire,
        )

        path = tmp_path / "weird.docx"
        path.write_text("not really a docx", encoding="utf-8")
        with pytest.raises(ValueError, match="Unsupported"):
            parse_completed_questionnaire(path)

    def test_missing_file_errors(self, tmp_path: Path) -> None:
        from evidentia_core.tprm.questionnaire import (
            parse_completed_questionnaire,
        )

        with pytest.raises(FileNotFoundError):
            parse_completed_questionnaire(tmp_path / "ghost.json")

    def test_parses_json_without_vendor_id_returns_none(
        self, tmp_path: Path
    ) -> None:
        """v0.7.9 P0.4 Continuous H-4: a questionnaire JSON whose
        prefill block carries no vendor_id (operator-edited file or
        vendor stripped the metadata) parses cleanly with
        vendor_id=None. Caller is responsible for surfacing a
        clear error if correlation isn't supplied via the CLI's
        --vendor-id override."""
        from evidentia_core.tprm.questionnaire import (
            parse_completed_questionnaire,
        )

        # Construct a minimal JSON questionnaire with empty vendor block
        json_path = tmp_path / "no-vendor-id.json"
        import json as _json

        json_path.write_text(
            _json.dumps(
                {
                    "id": "00000000-0000-0000-0000-000000000001",
                    "format": "evidentia-generic",
                    "questions": [
                        {
                            "id": "EVG-GOV-01",
                            "domain": "Governance",
                            "question_text": "?",
                            "vendor_response": "Yes",
                        }
                    ],
                    "vendor": {},  # empty — no vendor_id field
                }
            ),
            encoding="utf-8",
        )
        completed = parse_completed_questionnaire(json_path)
        assert completed.questionnaire_id == (
            "00000000-0000-0000-0000-000000000001"
        )
        assert completed.vendor_id is None
        assert completed.responses == {"EVG-GOV-01": "Yes"}


# ── v0.7.9 P0.2 second slice: SIG BYO template ─────────────────────


class TestGenerateFromByoTemplate:
    def _make_synthetic_sig(self, tmp_path: Any) -> Any:  # type: ignore[name-defined]
        """Build a synthetic XLSX shaped like a SIG layout (label-
        in-column-A, response-in-column-B)."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendor Information"
        ws.append(["Label", "Response"])
        ws.append(["Company Name", ""])
        ws.append(["Vendor Type", ""])
        ws.append(["Criticality Tier", ""])
        ws.append(["Primary Contact", ""])
        ws.append(["Contract Start Date", ""])
        ws.append(["Region", ""])
        # Add an unrelated questions sheet so the workbook is realistic
        q_ws = wb.create_sheet("Questions")
        q_ws.append(["QID", "Question", "Response"])
        q_ws.append(["Q-1", "Have you done X?", ""])
        path = tmp_path / "sig-template.xlsx"
        wb.save(str(path))
        return path

    def test_byo_sig_pre_fills_vendor_metadata(
        self, tmp_path: Any  # type: ignore[name-defined]
    ) -> None:
        import io

        import openpyxl
        from evidentia_core.tprm.questionnaire import (
            generate_from_byo_template,
        )

        template = self._make_synthetic_sig(tmp_path)
        v = _make_vendor(
            name="Acme Cloud",
            type_=VendorType.SAAS,
            criticality_tier=CriticalityTier.HIGH,
            region="us-east-1",
        )
        out = generate_from_byo_template(
            v, template_path=template, fmt=QuestionnaireFormat.SIG
        )
        assert isinstance(out, bytes)
        assert out[:2] == b"PK"  # XLSX zip magic
        wb = openpyxl.load_workbook(filename=io.BytesIO(out))
        ws = wb["Vendor Information"]
        rows = list(ws.iter_rows(values_only=True))
        # Find the company-name row and assert it was filled
        for label, response, *_ in rows:
            if label == "Company Name":
                assert response == "Acme Cloud"

    def test_byo_sig_refuses_non_byo_format(
        self, tmp_path: Any  # type: ignore[name-defined]
    ) -> None:
        from evidentia_core.tprm.questionnaire import (
            generate_from_byo_template,
        )

        template = self._make_synthetic_sig(tmp_path)
        v = _make_vendor()
        with pytest.raises(
            ValueError, match="does not accept a BYO template"
        ):
            generate_from_byo_template(
                v,
                template_path=template,
                fmt=QuestionnaireFormat.CAIQ_LITE,
            )

    def test_byo_sig_errors_on_missing_template(
        self, tmp_path: Any  # type: ignore[name-defined]
    ) -> None:
        from evidentia_core.tprm.questionnaire import (
            generate_from_byo_template,
        )

        v = _make_vendor()
        with pytest.raises(FileNotFoundError):
            generate_from_byo_template(
                v,
                template_path=tmp_path / "nonexistent.xlsx",
                fmt=QuestionnaireFormat.SIG,
            )

    def test_byo_sig_errors_on_unrecognized_layout(
        self, tmp_path: Path
    ) -> None:
        import openpyxl
        from evidentia_core.tprm.questionnaire import (
            generate_from_byo_template,
        )

        # XLSX with no recognizable vendor-metadata sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Random Sheet"
        ws.append(["Some", "Random", "Data"])
        path = tmp_path / "weird.xlsx"
        wb.save(str(path))
        v = _make_vendor()
        with pytest.raises(
            RuntimeError, match="No recognizable vendor-metadata"
        ):
            generate_from_byo_template(
                v, template_path=path, fmt=QuestionnaireFormat.SIG
            )

    def test_byo_sig_partial_label_match_succeeds(
        self, tmp_path: Path
    ) -> None:
        """v0.7.9 P0.4 Continuous H-4: a SIG template where SOME
        vendor-metadata labels match Evidentia's recognizer + others
        don't. The function should silently skip the non-matching
        rows and pre-fill the matching ones, NOT fail the whole
        operation."""
        import io

        import openpyxl
        from evidentia_core.tprm.questionnaire import (
            generate_from_byo_template,
        )

        # Mix of recognized + unrecognized labels. Note the layout:
        # column A label, column B empty, column C empty (3-column
        # rows so the H-5 prefer-C ordering applies; matched values
        # land in column C).
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendor Information"
        ws.append(["Label", "Instructions", "Response"])
        ws.append(["Some Custom Field", "instructions", ""])  # unmatched
        ws.append(["Company Name", "fill in legal name", ""])  # matched
        ws.append(["Internal SIG Q-99", "internal", ""])  # unmatched
        ws.append(["Vendor Type", "saas/iaas/etc", ""])  # matched
        path = tmp_path / "partial.xlsx"
        wb.save(str(path))

        v = _make_vendor(name="PartialMatch Co", type_=VendorType.SAAS)
        out = generate_from_byo_template(
            v, template_path=path, fmt=QuestionnaireFormat.SIG
        )
        # Should NOT raise; matching rows pre-filled, others left empty
        assert isinstance(out, bytes)
        wb_out = openpyxl.load_workbook(filename=io.BytesIO(out))
        ws_out = wb_out["Vendor Information"]
        rows = list(ws_out.iter_rows(values_only=True))
        # Build label→C-cell map for matched-rows assertion (H-5
        # prefers column C when present + empty)
        # row format: (label, instructions, response)
        for label, _instructions, response in rows[1:]:
            if label == "Company Name":
                assert response == "PartialMatch Co"
            if label == "Vendor Type":
                assert response == "saas"
            if label == "Some Custom Field":
                assert response in (None, "")
            if label == "Internal SIG Q-99":
                assert response in (None, "")
