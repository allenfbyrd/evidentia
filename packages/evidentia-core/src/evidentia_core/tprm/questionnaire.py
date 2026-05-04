"""Vendor due-diligence questionnaire generator (v0.7.9 P0.2).

Auto-generates SIG / SIG-Lite / CAIQ-style questionnaires for a
vendor in the inventory. Pre-fills vendor-metadata fields (name,
type, criticality tier, contract dates, regulatory classification,
4th-party disclosures) so the operator only sends the vendor
unanswered control questions, not blank templates.

Format catalogue (per ``QuestionnaireFormat`` enum):

- **``evidentia-generic``** — Apache 2.0 / Evidentia-original
  baseline questionnaire (~20 questions across FFIEC vendor-
  management domains: governance, access control, data handling,
  incident response, business continuity, 4th-party disclosure).

- **``caiq-lite``** — CC BY 4.0 / CSA CAIQ v4.0.3 representative
  ~25-question subset.

- **``caiq-full``** — CC BY 4.0 / CSA CAIQ v4.0.3 expanded
  ~50-question subset covering all 17 control domains. v0.7.9
  P0.2 second slice. Operators wanting the full 245-question
  CAIQ should download from
  <https://cloudsecurityalliance.org/research/cloud-controls-matrix/>
  and use the BYO ``--from-template`` path.

- **``sig``** / **``sig-lite``** — STUB content + BYO-template
  path (v0.7.9 P0.2 second slice). Shared Assessments paywalls
  SIG questions, but operators with a licensed XLSX can supply
  it via ``--from-template <path>``. Evidentia parses the SIG
  layout, pre-fills vendor metadata into standard cells, and
  returns the partially-filled XLSX for the vendor to complete.
  Without ``--from-template``, the format raises a clear
  "BYO-template required" error.

Output formats (per ``--output-format`` flag):

- **``json``** — full Pydantic model dump; machine-consumable
- **``csv``** — flat; spreadsheet-pivot friendly without binary deps
- **``xlsx``** — workbook with vendor-prefill header sheet +
  one questions sheet per domain. Requires
  ``pip install 'evidentia-core[xlsx]'`` (openpyxl).
  v0.7.9 P0.2 second slice.

Ingest path (v0.7.9 P0.2 second slice):

The ``ingest`` workflow loads a completed vendor questionnaire
back into Evidentia and correlates it to a vendor record via the
questionnaire's UUID (or an explicit ``--vendor-id`` override).
Supported input shapes: JSON / CSV / XLSX produced by the
generate path. Vendor evidence_refs[] is updated; subsequent
``evidentia tprm vendor show`` reflects the ingested response.
"""

from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime
from enum import Enum
from importlib import resources
from typing import Any

from pydantic import Field

from evidentia_core.models.common import (
    EvidentiaModel,
    current_version,
    new_id,
    utc_now,
)
from evidentia_core.models.tprm import Vendor
from evidentia_core.tprm.concentration import _csv_safe


class QuestionnaireFormat(str, Enum):
    """Supported questionnaire framework formats.

    The free-text values map to packaged data files under
    `evidentia_core.tprm.data.questionnaires.<value>.json`. Three
    of the five values are SHIPPED (``evidentia-generic``,
    ``caiq-lite``); two are stubs (``sig``, ``sig-lite``) that
    document the future BYO-template path but error today.
    """

    EVIDENTIA_GENERIC = "evidentia-generic"
    CAIQ_LITE = "caiq-lite"
    CAIQ_FULL = "caiq-full"
    SIG = "sig"
    SIG_LITE = "sig-lite"


# Format → data-file basename mapping. Stubs (sig/sig-lite) have
# no entry and raise NotImplementedError at generate time UNLESS
# a ``--from-template`` path is supplied (BYO XLSX ingestion).
_PACKAGED_DATA_FORMATS: dict[QuestionnaireFormat, str] = {
    QuestionnaireFormat.EVIDENTIA_GENERIC: "evidentia_generic.json",
    QuestionnaireFormat.CAIQ_LITE: "caiq_lite.json",
    QuestionnaireFormat.CAIQ_FULL: "caiq_full.json",
}

# SIG / SIG-Lite formats that accept BYO XLSX templates (no
# packaged data; content comes from the operator's licensed file).
_BYO_TEMPLATE_FORMATS: frozenset[QuestionnaireFormat] = frozenset(
    {QuestionnaireFormat.SIG, QuestionnaireFormat.SIG_LITE}
)


class Question(EvidentiaModel):
    """One control question in a vendor due-diligence questionnaire."""

    id: str = Field(
        description=(
            "Stable per-format identifier. Format-prefixed (e.g., "
            "'EVG-GOV-01' for evidentia-generic governance question 1, "
            "'CAIQ-AAC-01' for CAIQ Audit Assurance & Compliance #1). "
            "Pinned so operator responses can flow back via the planned "
            "ingest command without column-order drift."
        )
    )
    domain: str = Field(
        description=(
            "Framework-specific control domain (e.g., 'Governance', "
            "'Access Control', 'Data Handling' for evidentia-generic; "
            "'AAC' (Audit Assurance & Compliance), 'BCR' (Business "
            "Continuity & Resilience) for CAIQ)."
        )
    )
    question_text: str = Field(
        description=(
            "Verbatim question presented to the vendor. Authoritative "
            "wording for caiq-lite per CSA CAIQ v4.0.3 publications "
            "(CC BY 4.0); Evidentia-original wording for "
            "evidentia-generic (Apache 2.0)."
        )
    )
    response_options: list[str] = Field(
        default_factory=list,
        description=(
            "Optional structured response choices (e.g., "
            "['Yes', 'No', 'Partial', 'Not Applicable']). Empty list "
            "= free-text response expected."
        ),
    )
    notes: str | None = Field(
        default=None,
        description=(
            "Optional clarifier surfaced to the vendor (e.g., 'Provide "
            "your most recent SOC 2 Type II report period covered')."
        ),
    )


class VendorPreFill(EvidentiaModel):
    """Pre-filled vendor metadata included in the questionnaire output.

    Operators DON'T want to ask the vendor 'what's your legal name?'
    when Evidentia already has it. This block carries the
    questionnaire-time snapshot of the vendor record so the receiving
    party sees the operator's understanding + can flag corrections.
    """

    vendor_id: str
    vendor_name: str
    vendor_type: str
    criticality_tier: str
    relationship_owner: str
    contract_start_date: date
    contract_end_date: date | None = None
    region: str | None = None
    regulatory_classification: list[str] = Field(default_factory=list)
    fourth_party_count: int = 0
    fourth_party_names: list[str] = Field(default_factory=list)


class Questionnaire(EvidentiaModel):
    """Generated vendor due-diligence questionnaire.

    Round-trip: serialise via ``model_dump_json(indent=2)`` to ship
    to the vendor as a structured doc; or render via
    :func:`render_csv_questionnaire` for spreadsheet workflows. The
    ``id`` field carries a UUID so a future ``dd-questionnaire ingest``
    command can correlate responses back to the originating
    questionnaire.
    """

    id: str = Field(default_factory=new_id)
    format: QuestionnaireFormat
    title: str = Field(
        description=(
            "Human-readable header — e.g., 'Evidentia Generic Vendor "
            "DD Questionnaire — <vendor name>'."
        )
    )
    generated_at: datetime = Field(default_factory=utc_now)
    vendor: VendorPreFill = Field(
        description="Snapshot of vendor metadata pre-filled into the questionnaire."
    )
    questions: list[Question] = Field(default_factory=list)
    licensing_attribution: str | None = Field(
        default=None,
        description=(
            "Format-specific attribution required by the source "
            "license (e.g., CSA CC BY 4.0 attribution for caiq-lite)."
        ),
    )
    evidentia_version: str = Field(default_factory=current_version)


# ── data loading ──────────────────────────────────────────────────


def _load_questionnaire_data(fmt: QuestionnaireFormat) -> dict[str, Any]:
    """Load packaged JSON data for a supported format.

    Raises :class:`NotImplementedError` for stub formats
    (``sig`` / ``sig-lite``) until the BYO-template path lands.
    """
    if fmt not in _PACKAGED_DATA_FORMATS:
        raise NotImplementedError(
            f"Questionnaire format {fmt.value!r} is not yet shipped. "
            "SIG / SIG-Lite require a Shared Assessments licensed "
            "template (paywalled content); future versions will "
            "support `--from-template <licensed-xlsx>` BYO ingestion. "
            f"Currently shipped formats: "
            f"{sorted(f.value for f in _PACKAGED_DATA_FORMATS)}."
        )
    fname = _PACKAGED_DATA_FORMATS[fmt]
    with resources.files(
        "evidentia_core.tprm.data.questionnaires"
    ).joinpath(fname).open("r", encoding="utf-8") as fp:
        data = json.load(fp)
    return data  # type: ignore[no-any-return]


# ── prefill ───────────────────────────────────────────────────────


def _build_prefill(vendor: Vendor) -> VendorPreFill:
    """Project a Vendor into the questionnaire's prefill block."""
    return VendorPreFill(
        vendor_id=vendor.id,
        vendor_name=vendor.name,
        vendor_type=str(vendor.type),
        criticality_tier=str(vendor.criticality_tier),
        relationship_owner=vendor.relationship_owner,
        contract_start_date=vendor.contract_start_date,
        contract_end_date=vendor.contract_end_date,
        region=vendor.region,
        regulatory_classification=[
            str(c) for c in vendor.regulatory_classification
        ],
        fourth_party_count=len(vendor.fourth_parties),
        fourth_party_names=[fp.name for fp in vendor.fourth_parties],
    )


# ── core ──────────────────────────────────────────────────────────


def generate_questionnaire(
    vendor: Vendor,
    format: QuestionnaireFormat,
) -> Questionnaire:
    """Generate a due-diligence questionnaire for ``vendor`` in ``format``.

    Args:
        vendor: The vendor inventory record.
        format: One of the :class:`QuestionnaireFormat` values.

    Returns:
        A populated :class:`Questionnaire`.

    Raises:
        NotImplementedError: format is a stub (sig / sig-lite).
    """
    data = _load_questionnaire_data(format)
    questions = [Question.model_validate(q) for q in data["questions"]]
    title_template = data.get(
        "title_template", "Vendor DD Questionnaire — {vendor_name}"
    )
    # Use string `.replace()` rather than `.format(vendor_name=...)`
    # so a vendor name containing `{}` / `{0}` / `{secret}` cannot
    # raise KeyError at render time or, worse, leak adjacent format
    # args via positional/attribute walking. Closes v0.7.9 P0.3+P0.2
    # Continuous-review H-3 (defensive). The template comes from
    # packaged trusted JSON, so it itself can't be malicious — the
    # concern is just that vendor.name is user-supplied free-text.
    title = title_template.replace("{vendor_name}", vendor.name)
    return Questionnaire(
        format=format,
        title=title,
        vendor=_build_prefill(vendor),
        questions=questions,
        licensing_attribution=data.get("licensing_attribution"),
    )


# ── rendering ─────────────────────────────────────────────────────


def render_csv_questionnaire(q: Questionnaire) -> str:
    """Render a Questionnaire as flat CSV.

    Header row carries the prefill vendor metadata (one column per
    field) so the operator can confirm at a glance that the vendor
    they're asking matches their understanding. The question rows
    follow with columns: id / domain / question_text / response_options
    / notes / vendor_response (blank for the vendor to complete).
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    # Header section: vendor prefill as key/value comment lines.
    # Use a leading "#" sentinel so spreadsheet importers can filter
    # them out if desired; csv module handles arbitrary strings fine.
    # User-content cells go through _csv_safe to neutralize formula-
    # injection vectors (vendor name + 4th-party name + region etc.
    # are operator-supplied; the questionnaire CSV is explicitly
    # designed to be sent to the vendor, making formula injection
    # in those fields a real exfil/phish primitive). Closes v0.7.9
    # P0.2/P0.3 Continuous-review H-1 / security M-1.
    writer.writerow(["# Vendor DD Questionnaire"])
    writer.writerow(["# Title", _csv_safe(q.title)])
    writer.writerow(["# Questionnaire ID", q.id])
    writer.writerow(["# Format", str(q.format)])
    writer.writerow(["# Generated at", q.generated_at.isoformat()])
    writer.writerow(["# Vendor ID", q.vendor.vendor_id])
    writer.writerow(["# Vendor name", _csv_safe(q.vendor.vendor_name)])
    writer.writerow(["# Vendor type", q.vendor.vendor_type])
    writer.writerow(["# Criticality tier", q.vendor.criticality_tier])
    writer.writerow(
        ["# Relationship owner", _csv_safe(q.vendor.relationship_owner)]
    )
    writer.writerow(
        ["# Contract start", str(q.vendor.contract_start_date)]
    )
    if q.vendor.contract_end_date:
        writer.writerow(
            ["# Contract end", str(q.vendor.contract_end_date)]
        )
    if q.vendor.region:
        writer.writerow(["# Region", _csv_safe(q.vendor.region)])
    if q.vendor.regulatory_classification:
        writer.writerow(
            [
                "# Regulatory classification",
                _csv_safe(", ".join(q.vendor.regulatory_classification)),
            ]
        )
    if q.vendor.fourth_party_count:
        writer.writerow(
            [
                "# 4th parties",
                _csv_safe(
                    f"{q.vendor.fourth_party_count}: "
                    + ", ".join(q.vendor.fourth_party_names)
                ),
            ]
        )
    if q.licensing_attribution:
        writer.writerow(["# Attribution", q.licensing_attribution])
    writer.writerow([])  # blank separator row
    # Question rows. Question content (text + notes) comes from
    # packaged-data trusted JSON, but we _csv_safe it anyway as
    # defense-in-depth against future user-templated questions.
    writer.writerow(
        [
            "id",
            "domain",
            "question_text",
            "response_options",
            "notes",
            "vendor_response",
        ]
    )
    for question in q.questions:
        writer.writerow(
            [
                question.id,
                question.domain,
                _csv_safe(question.question_text),
                " | ".join(question.response_options),
                _csv_safe(question.notes) if question.notes else "",
                "",  # blank for vendor to fill
            ]
        )
    return buf.getvalue()


# ── catalogue ─────────────────────────────────────────────────────


def shipped_formats() -> list[QuestionnaireFormat]:
    """Return the sorted list of formats with packaged data.

    Used by CLI + REST help text to enumerate valid choices without
    hard-coding the stub-vs-shipped split in three places.
    """
    return sorted(_PACKAGED_DATA_FORMATS.keys(), key=lambda f: f.value)


def byo_template_formats() -> list[QuestionnaireFormat]:
    """Return the sorted list of formats that accept BYO XLSX templates.

    Used by CLI help text to differentiate between packaged-content
    formats (caiq-lite / caiq-full / evidentia-generic) and BYO-only
    formats (sig / sig-lite).
    """
    return sorted(_BYO_TEMPLATE_FORMATS, key=lambda f: f.value)


# ── v0.7.9 P0.2 second slice: XLSX render ─────────────────────────


class XlsxNotInstalledError(ImportError):
    """Raised when XLSX functionality is requested but openpyxl is missing.

    The operator can resolve by installing the optional extra::

        pip install 'evidentia-core[xlsx]'
    """


def _require_openpyxl() -> Any:
    """Lazy-import openpyxl with a clear, actionable error message."""
    try:
        import openpyxl  # type: ignore[import-untyped, unused-ignore]
    except ImportError as e:
        raise XlsxNotInstalledError(
            "openpyxl is required for XLSX output / ingestion. "
            "Install via `pip install 'evidentia-core[xlsx]'`. "
            "openpyxl is ~3 MB pure-Python; gated behind an extra "
            "to keep default installs slim."
        ) from e
    return openpyxl


def render_xlsx_questionnaire(q: Questionnaire) -> bytes:
    """Render a Questionnaire as an Excel workbook (.xlsx) bytes.

    Layout:
    - Sheet 1: ``Vendor metadata`` — all VendorPreFill fields as
      key/value pairs, plus questionnaire ID + format + generated_at.
    - Sheet 2..N: one sheet per question domain (sanitized to fit
      Excel's 31-character sheet-name limit + reserved chars). Each
      domain sheet has columns: id / question / response_options /
      notes / vendor_response (blank for the vendor to complete).

    The XLSX file inherits CSV-formula-injection defenses for
    user-supplied content cells: vendor.name / fourth_party_names /
    region / relationship_owner all flow through the same
    `_csv_safe()` neutralizer used by the CSV path. The XLSX file
    format does NOT auto-evaluate formulas the same way CSV
    spreadsheet importers do, but defense-in-depth is cheap.

    Args:
        q: The generated questionnaire to render.

    Returns:
        XLSX file content as bytes (ready to write to disk or
        return via REST). The caller is responsible for I/O.

    Raises:
        XlsxNotInstalledError: openpyxl is not installed; install
            the ``[xlsx]`` extra.
    """
    openpyxl = _require_openpyxl()
    wb = openpyxl.Workbook()
    # Default sheet becomes the metadata sheet
    meta_ws = wb.active
    meta_ws.title = "Vendor metadata"
    meta_rows: list[list[str]] = [
        ["Title", _csv_safe(q.title)],
        ["Questionnaire ID", q.id],
        ["Format", str(q.format)],
        ["Generated at", q.generated_at.isoformat()],
        ["Vendor ID", q.vendor.vendor_id],
        ["Vendor name", _csv_safe(q.vendor.vendor_name)],
        ["Vendor type", q.vendor.vendor_type],
        ["Criticality tier", q.vendor.criticality_tier],
        ["Relationship owner", _csv_safe(q.vendor.relationship_owner)],
        ["Contract start", str(q.vendor.contract_start_date)],
    ]
    if q.vendor.contract_end_date:
        meta_rows.append(["Contract end", str(q.vendor.contract_end_date)])
    if q.vendor.region:
        meta_rows.append(["Region", _csv_safe(q.vendor.region)])
    if q.vendor.regulatory_classification:
        meta_rows.append(
            [
                "Regulatory classification",
                _csv_safe(", ".join(q.vendor.regulatory_classification)),
            ]
        )
    if q.vendor.fourth_party_count:
        meta_rows.append(
            [
                "4th parties",
                _csv_safe(
                    f"{q.vendor.fourth_party_count}: "
                    + ", ".join(q.vendor.fourth_party_names)
                ),
            ]
        )
    if q.licensing_attribution:
        meta_rows.append(["Attribution", q.licensing_attribution])
    for row in meta_rows:
        meta_ws.append(row)

    # Group questions by domain → one sheet per domain.
    by_domain: dict[str, list[Question]] = {}
    for question in q.questions:
        by_domain.setdefault(question.domain, []).append(question)

    for domain, questions in by_domain.items():
        sheet_name = _sanitize_sheet_name(domain)
        # Excel requires unique sheet names; the sanitizer does its
        # best but collisions can still happen if two domains
        # truncate identically.
        suffix = ""
        attempt = 0
        while (sheet_name + suffix) in wb.sheetnames:
            attempt += 1
            suffix = f" ({attempt})"
        ws = wb.create_sheet(title=sheet_name + suffix)
        ws.append(
            [
                "id",
                "question_text",
                "response_options",
                "notes",
                "vendor_response",
            ]
        )
        for question in questions:
            ws.append(
                [
                    question.id,
                    _csv_safe(question.question_text),
                    " | ".join(question.response_options),
                    _csv_safe(question.notes) if question.notes else "",
                    "",  # blank for vendor to complete
                ]
            )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# Excel sheet-name constraints (OOXML spec + legacy quirks):
# - Max 31 characters
# - Cannot contain: : \ / ? * [ ]
# - Cannot start or end with a single quote
# - v0.7.10 P3 closure of v0.7.9 L-3: tilde (~) added for
#   defense-in-depth — OOXML doesn't reserve it but legacy
#   Excel-on-Mac variants flag it as a workbook-name conflict
#   character; cheaper to strip than to debug an auditor's
#   "Excel doesn't open this" report.
_EXCEL_SHEET_BAD_CHARS = ":\\/?*[]~"


def _sanitize_sheet_name(name: str, *, reserve: int = 4) -> str:
    """Sanitize a domain string into a valid Excel sheet name.

    Excel's hard limit is 31 characters per sheet name. Callers that
    apply collision-suffixes to deduplicate sheet names (e.g.,
    ``" (2)"``, ``" (3)"``, ``" (10)"``) need to reserve room for the
    suffix. The default ``reserve=4`` covers up to ``" (99)"`` —
    sufficient for the typical TPRM portfolio sizes.

    v0.7.12 P3 closure of v0.7.9 M-7: previously truncated to 31
    chars unconditionally, leaving callers no room for collision
    suffixes. The new contract truncates to ``31 - reserve``,
    preserving the canonical 31-char ceiling AFTER suffix
    application.
    """
    cleaned = "".join(
        c for c in name if c not in _EXCEL_SHEET_BAD_CHARS
    )
    cleaned = cleaned.strip().strip("'")
    if not cleaned:
        cleaned = "Questions"
    max_base = max(1, 31 - max(0, reserve))
    return cleaned[:max_base]


# ── v0.7.9 P0.2 second slice: SIG BYO template ────────────────────


def generate_from_byo_template(
    vendor: Vendor,
    *,
    template_path: Any,
    fmt: QuestionnaireFormat,
) -> bytes:
    """Pre-fill vendor metadata into an operator-supplied SIG XLSX.

    Operators with Shared Assessments membership supply their
    licensed SIG / SIG-Lite XLSX template. Evidentia opens the
    workbook, locates the standard "Vendor Information" / "Company
    Information" sheet, pre-fills vendor metadata into the
    documented cells, and returns the partially-filled file as bytes
    for the vendor to complete the control questions.

    The SIG template's question content stays UNTOUCHED — Evidentia
    only writes to vendor-metadata cells. This respects Shared
    Assessments' license terms (no redistribution of question
    content; operator's licensed copy stays on the operator's
    machine).

    Cell-coordinate convention (best-effort, defensive):
    The SIG / SIG-Lite layouts vary across Shared Assessments
    publication years. The function looks for a sheet whose name
    contains "vendor information" / "company information" /
    "company profile" (case-insensitive), then writes to the FIRST
    cell in column B / C of rows whose column-A label matches
    well-known SIG vendor-metadata labels (Company Name / Vendor
    Name / Legal Name / Address / Primary Contact / etc.).

    If no recognized layout is detected, the function raises a
    clear error with debug info and the operator can fall back to
    the JSON / CSV / XLSX evidentia-generic / caiq-lite formats
    (which carry their own layouts).

    Args:
        vendor: The vendor to pre-fill metadata from.
        template_path: Path to the operator's licensed SIG /
            SIG-Lite XLSX file.
        fmt: Either ``QuestionnaireFormat.SIG`` or
            ``QuestionnaireFormat.SIG_LITE``. Used in error
            messages only — both follow the same layout convention.

    Returns:
        Pre-filled XLSX bytes ready for the operator to send the
        vendor.

    Raises:
        XlsxNotInstalledError: openpyxl missing.
        ValueError: ``fmt`` is not a BYO-template format.
        FileNotFoundError: template_path does not exist or isn't
            readable.
        RuntimeError: no recognized SIG layout found in the workbook.
    """
    if fmt not in _BYO_TEMPLATE_FORMATS:
        raise ValueError(
            f"Format {fmt.value!r} does not accept a BYO template. "
            "Only sig / sig-lite use --from-template."
        )
    openpyxl = _require_openpyxl()
    from pathlib import Path as _Path

    path = _Path(str(template_path))
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(
            f"BYO SIG template not found at {path!r}. Operators "
            "with Shared Assessments membership can download the "
            "current SIG / SIG-Lite XLSX from "
            "https://sharedassessments.org/sig/ and supply the "
            "path here."
        )
    wb = openpyxl.load_workbook(filename=str(path), data_only=False)

    # Find the vendor-metadata sheet using fuzzy name matching.
    target_sheet_name: str | None = None
    name_keywords = (
        "vendor information",
        "company information",
        "company profile",
        "vendor profile",
        "general information",
    )
    for sheet_name in wb.sheetnames:
        normalized = sheet_name.lower().strip()
        if any(kw in normalized for kw in name_keywords):
            target_sheet_name = sheet_name
            break
    if target_sheet_name is None:
        raise RuntimeError(
            f"No recognizable vendor-metadata sheet found in "
            f"{path.name!r}. Expected a sheet named one of: "
            f"{', '.join(name_keywords)}. Workbook sheets: "
            f"{wb.sheetnames}. The SIG layout has changed over "
            "time; operators may need to manually populate vendor "
            "metadata in the licensed template."
        )
    ws = wb[target_sheet_name]

    # Map column-A labels (lowercased + normalized) to vendor values.
    # Defensive: if an operator's template uses different labels,
    # we silently skip the fields that don't match — better partial
    # pre-fill than failure.
    label_to_value: dict[str, str] = {
        "company name": vendor.name,
        "vendor name": vendor.name,
        "legal name": vendor.name,
        "vendor": vendor.name,
        "primary contact": vendor.relationship_owner,
        "primary contact email": vendor.relationship_owner,
        "relationship owner": vendor.relationship_owner,
        "contract start date": vendor.contract_start_date.isoformat(),
        "contract effective date": vendor.contract_start_date.isoformat(),
        "vendor type": (
            vendor.type.value
            if hasattr(vendor.type, "value")
            else str(vendor.type)
        ),
        "service type": (
            vendor.type.value
            if hasattr(vendor.type, "value")
            else str(vendor.type)
        ),
        "criticality": (
            vendor.criticality_tier.value
            if hasattr(vendor.criticality_tier, "value")
            else str(vendor.criticality_tier)
        ),
        "criticality tier": (
            vendor.criticality_tier.value
            if hasattr(vendor.criticality_tier, "value")
            else str(vendor.criticality_tier)
        ),
        "tier": (
            vendor.criticality_tier.value
            if hasattr(vendor.criticality_tier, "value")
            else str(vendor.criticality_tier)
        ),
    }
    if vendor.contract_end_date:
        label_to_value["contract end date"] = (
            vendor.contract_end_date.isoformat()
        )
        label_to_value["contract expiration date"] = (
            vendor.contract_end_date.isoformat()
        )
    if vendor.region:
        label_to_value["region"] = vendor.region
        label_to_value["geographic region"] = vendor.region

    # Walk rows; for each label-match, write value into column B
    # (most common SIG layout) AND column C (some templates put
    # response in column C with column B as instructions). We
    # check the cell is empty first so we don't clobber existing
    # operator content.
    pre_filled_count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if not row:
            continue
        a_cell = row[0]
        if a_cell.value is None:
            continue
        a_value = str(a_cell.value).strip().lower().rstrip(":")
        target = label_to_value.get(a_value)
        if target is None:
            continue
        # v0.7.9 P0.4 Continuous H-5: real-world Shared Assessments
        # SIG / SIG-Lite templates frequently put instruction text in
        # column B and intend column C as the vendor response cell.
        # If column C exists and is empty, prefer it. Only fall back
        # to column B when column C is absent or already populated.
        # This produces correct SIG-layout results far more often than
        # the previous always-prefer-B order.
        if len(row) > 2 and row[2].value in (None, ""):
            row[2].value = target
            pre_filled_count += 1
        elif len(row) > 1 and row[1].value in (None, ""):
            row[1].value = target
            pre_filled_count += 1

    if pre_filled_count == 0:
        raise RuntimeError(
            f"No SIG vendor-metadata cells matched the pre-fill "
            f"label list in {path.name!r}. The template's row "
            "labels may differ from the documented Shared "
            "Assessments convention; operators may need to "
            "populate metadata manually. Recognized labels: "
            f"{sorted(label_to_value.keys())}."
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── v0.7.9 P0.2 second slice: ingest path ─────────────────────────


class CompletedQuestionnaire(EvidentiaModel):
    """A completed (or partially-completed) vendor questionnaire response.

    Produced by :func:`parse_completed_questionnaire`. Carries the
    same questionnaire UUID as the originating Questionnaire so the
    ingest CLI can correlate back without ambiguity. The
    ``responses`` mapping is keyed by question.id (e.g.,
    ``EVG-GOV-01`` / ``CAIQ-AAC-01``).
    """

    questionnaire_id: str | None = Field(
        default=None,
        description=(
            "UUID from the originating Questionnaire. Required for "
            "automatic vendor correlation; can be omitted if the "
            "ingest call passes --vendor-id explicitly."
        ),
    )
    vendor_id: str | None = Field(
        default=None,
        description="Vendor ID from the originating prefill (when present).",
    )
    format: QuestionnaireFormat | None = None
    responses: dict[str, str] = Field(
        default_factory=dict,
        description=(
            "Per-question vendor responses keyed by question.id. "
            "Empty string == 'no response'. Operators can post-"
            "process this map for compliance scoring or evidence-"
            "of-due-diligence claims."
        ),
    )
    ingested_at: datetime = Field(default_factory=utc_now)
    source_path: str | None = None


def parse_completed_questionnaire(
    path: Any,
) -> CompletedQuestionnaire:
    """Parse a vendor's completed questionnaire from disk.

    Auto-detects format from the file extension:

    - ``.json`` — uses the canonical Questionnaire shape; pulls
      ``vendor_response`` per question (when present), plus
      ``vendor_id`` and questionnaire ``id`` for correlation
    - ``.csv`` — reads the flat CSV produced by
      :func:`render_csv_questionnaire`; correlates via the
      ``# Questionnaire ID`` and ``# Vendor ID`` header rows
    - ``.xlsx`` — reads workbooks produced by
      :func:`render_xlsx_questionnaire`; correlates via
      "Vendor metadata" sheet's ``Questionnaire ID`` / ``Vendor ID``
      rows; question responses come from per-domain sheets'
      ``vendor_response`` column

    Args:
        path: Filesystem path to the completed questionnaire.

    Returns:
        :class:`CompletedQuestionnaire` carrying the responses.

    Raises:
        FileNotFoundError: path doesn't exist.
        ValueError: extension not supported, or malformed content.
        XlsxNotInstalledError: ``.xlsx`` requested without openpyxl.
    """
    from pathlib import Path as _Path

    p = _Path(str(path))
    if not p.exists():
        raise FileNotFoundError(f"Questionnaire file not found: {p!r}")
    suffix = p.suffix.lower()
    if suffix == ".json":
        return _parse_completed_json(p)
    if suffix == ".csv":
        return _parse_completed_csv(p)
    if suffix == ".xlsx":
        return _parse_completed_xlsx(p)
    raise ValueError(
        f"Unsupported questionnaire file extension: {suffix!r}. "
        "Supported: .json, .csv, .xlsx."
    )


def _parse_completed_json(path: Any) -> CompletedQuestionnaire:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(
            "Completed questionnaire JSON must be an object "
            "(matching the Questionnaire shape)."
        )
    responses: dict[str, str] = {}
    for q in raw.get("questions", []):
        if not isinstance(q, dict):
            continue
        qid = q.get("id")
        resp = q.get("vendor_response", "")
        if isinstance(qid, str) and qid:
            responses[qid] = str(resp) if resp is not None else ""
    fmt_raw = raw.get("format")
    fmt: QuestionnaireFormat | None = None
    if isinstance(fmt_raw, str):
        try:
            fmt = QuestionnaireFormat(fmt_raw)
        except ValueError:
            fmt = None
    vendor_block = raw.get("vendor") or {}
    vendor_id_raw = (
        vendor_block.get("vendor_id")
        if isinstance(vendor_block, dict)
        else None
    )
    return CompletedQuestionnaire(
        questionnaire_id=raw.get("id") if isinstance(raw.get("id"), str) else None,
        vendor_id=vendor_id_raw if isinstance(vendor_id_raw, str) else None,
        format=fmt,
        responses=responses,
        source_path=str(path),
    )


def _parse_completed_csv(path: Any) -> CompletedQuestionnaire:
    """Parse the flat CSV emitted by render_csv_questionnaire.

    Header rows start with '#' sentinel; question rows are tabular
    with header columns id / domain / question_text / response_options
    / notes / vendor_response. The vendor_response column is what
    we extract.
    """
    text = path.read_text(encoding="utf-8")
    reader = csv.reader(io.StringIO(text))
    questionnaire_id: str | None = None
    vendor_id: str | None = None
    fmt: QuestionnaireFormat | None = None
    headers: list[str] | None = None
    responses: dict[str, str] = {}
    for row in reader:
        if not row:
            continue
        first = row[0].strip()
        if first.startswith("#"):
            # Header / metadata row
            label = first.lstrip("#").strip().lower()
            value = row[1].strip() if len(row) > 1 else ""
            if label == "questionnaire id":
                questionnaire_id = value or None
            elif label == "vendor id":
                vendor_id = value or None
            elif label == "format":
                try:
                    fmt = QuestionnaireFormat(value)
                except ValueError:
                    fmt = None
            continue
        if headers is None:
            # First non-comment row is the column header
            headers = [c.strip() for c in row]
            continue
        # Question response row
        col_map = dict(zip(headers, row, strict=False))
        qid = col_map.get("id", "").strip()
        if not qid:
            continue
        responses[qid] = col_map.get("vendor_response", "").strip()
    return CompletedQuestionnaire(
        questionnaire_id=questionnaire_id,
        vendor_id=vendor_id,
        format=fmt,
        responses=responses,
        source_path=str(path),
    )


_METADATA_SHEET_CANDIDATES = (
    "Vendor metadata",
    "Vendor Metadata",
    "Metadata",
    "Vendor Information",
    "Cover",
)


def _parse_completed_xlsx(path: Any) -> CompletedQuestionnaire:
    """Parse XLSX workbooks emitted by render_xlsx_questionnaire.

    v0.7.12 P3 closure of v0.7.9 M-8: previously hard-coded the
    metadata sheet name as ``"Vendor metadata"``. The new contract
    accepts the canonical name plus 4 common variants (case + word
    order + cover-style aliases) so an operator who manually
    renames the sheet during their review pass doesn't silently lose
    metadata round-trip.
    """
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(filename=str(path), data_only=True)
    questionnaire_id: str | None = None
    vendor_id: str | None = None
    fmt: QuestionnaireFormat | None = None
    metadata_sheet_name: str | None = next(
        (name for name in _METADATA_SHEET_CANDIDATES if name in wb.sheetnames),
        None,
    )
    if metadata_sheet_name is not None:
        ws = wb[metadata_sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip().lower()
            value = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if label == "questionnaire id":
                questionnaire_id = value or None
            elif label == "vendor id":
                vendor_id = value or None
            elif label == "format":
                try:
                    fmt = QuestionnaireFormat(value)
                except ValueError:
                    fmt = None
    responses: dict[str, str] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == metadata_sheet_name:
            continue
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        if not header_row:
            continue
        headers = [str(c).strip() if c else "" for c in header_row]
        try:
            id_idx = headers.index("id")
            resp_idx = headers.index("vendor_response")
        except ValueError:
            continue
        for row in rows_iter:
            if not row or row[id_idx] is None:
                continue
            qid = str(row[id_idx]).strip()
            resp_cell = (
                row[resp_idx] if resp_idx < len(row) else None
            )
            responses[qid] = (
                str(resp_cell).strip() if resp_cell is not None else ""
            )
    return CompletedQuestionnaire(
        questionnaire_id=questionnaire_id,
        vendor_id=vendor_id,
        format=fmt,
        responses=responses,
        source_path=str(path),
    )
